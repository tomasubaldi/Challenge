from RPA.Browser.Selenium import Selenium
from config import Config
from info import Info
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import time

class NewYorkTimes:

    def __init__(self):
        self.browser= Selenium(auto_close=False)  
        self.url=Config().NEW_YORK_TIMES_URL

    def open_website(self):
        self.browser.open_available_browser(self.url)
        self.browser.maximize_browser_window()

    def search_phrase(self, search_phrase=Config().SEARCH_PHRASE):
        self.browser.click_button_when_visible("//button[@class='css-tkwi90 e1iflr850']")
        self.browser.input_text_when_element_is_visible("//input[@placeholder='SEARCH']",search_phrase)
        self.browser.click_button_when_visible("//button[normalize-space()='Go']")
        


    def apply_filters(self, sections=Config().SECTION):
        self.browser.select_from_list_by_value("css=.css-v7it2b", "newest")
        self.browser.click_button_when_visible("//div[@data-testid='section']//button[@type='button']")
        self.browser.wait_until_element_is_visible("//ul[@class='css-64f9ga']")
        section_elements = self.browser.find_elements("//ul[@class='css-64f9ga']//span[@class='css-16eo56s']")
        time.sleep(2)
        for section_element in section_elements:
            for section in sections:
                if section in section_element.text:
                    section_element.click()
        self.browser.click_button_when_visible("//div[@data-testid='section']//button[@type='button']")
        while self.browser.is_element_visible("//button[normalize-space()='Show More']"):
            self.browser.click_button_when_visible("//button[normalize-space()='Show More']")
            time.sleep(1)

       
    def _get_title(self, new: Info, i:int):
        try:
            title_element = self.browser.find_element(f"(//h4[@class='css-2fgx4k'])[{i}]")
            new.title = title_element.text 
        except: 
            pass
    
    
    def _get_description(self, new: Info, i:int):
        try:
            description_element = self.browser.find_element(f"(//p[@class='css-16nhkrn'])[{i}]")
            new.description = description_element.text 
        except: 
            pass
    
    
    def _get_date(self, new: Info, i:int):
        try:
            date_element = self.browser.find_element(f"(//span[@class='css-17ubb9w'])[{i}]")
            new.date = date_element.text 
        except: 
            pass
    
    def _is_a_new(self, i):
        try:
            title_element = self.browser.find_element(f"(//h4[@class='css-2fgx4k'])[{i}]")
            if title_element.text == "":
                return False
            else:
                return True
        except Exception as ex: 
            raise ex
    
    
    def get_news_data(self):
        self.browser.wait_until_element_is_visible("//div[@class='css-1wa7u5r']")
        news_list = []
        time.sleep(2)
        news_elements = self.browser.find_elements("//main[@id='site-content']//li")
        for i, _ in enumerate(news_elements,start=1):
            try:
                if self._is_a_new(i):
                    new_info = Info()
                    self._get_title(new_info, i)
                    self._get_date(new_info, i)
                    self._get_description(new_info, i)
                    #self._count_search_phrases(new_info)
                    news_list.append(new_info)
            except:
                continue
        return news_list

    def save_to_excel(self, news_list):
        wb = openpyxl.Workbook()
        ws = wb.active
  
        ws["A1"] = "Title"
        ws["B1"] = "Date"
        ws["C1"] = "Description"
        ws["D1"] = "Image Filename"
        ws["E1"] = "Search Phrase Count"
        ws["F1"] = "Contains Money"


        for i, news in enumerate(news_list, start=2):
            ws[f"A{i}"] = news.title
            ws[f"B{i}"] = news.date
            ws[f"C{i}"] = news.description
            #ws[f"D{i}"] = news.image_filename
            ws[f"E{i}"] = news.search_phrase_count
            #ws[f"F{i}"] = news.("contains_money", "")

        
        wb.save("news.xlsx")






        


