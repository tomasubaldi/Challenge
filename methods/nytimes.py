from RPA.Browser.Selenium import Selenium
from config import Config
from info import Info
import openpyxl
import time
import datetime
from dateutil.relativedelta import relativedelta
from selenium.webdriver.common.keys import Keys
import re
import urllib.request
import os 
from openpyxl.styles import Font


class NewYorkTimes:

    def __init__(self):
        self.browser= Selenium(auto_close=False)  
        self.url=Config().NEW_YORK_TIMES_URL
        self.cwd = os.getcwd()
    
    #Method to Open Website and check the cookies banner
    def open_website(self):
        self.browser.open_available_browser(self.url)
        self.browser.maximize_browser_window()
        #If the Cookies Banner Appears
        if self.browser.is_element_visible('//*[@id="site-content"]/div[2]/div[2]/div/div[2]/button[1]'):
           self.browser.click_button_when_visible('//*[@id="site-content"]/div[2]/div[2]/div/div[2]/button[1]')
    
    #Search Phrase Method
    def search_phrase(self, search_phrase=Config().SEARCH_PHRASE):
        print("Searching phrase...")
        self.browser.click_button_when_visible("//button[@class='css-tkwi90 e1iflr850']")
        self.browser.input_text_when_element_is_visible("//input[@placeholder='SEARCH']",search_phrase)
        self.browser.click_button_when_visible("//button[normalize-space()='Go']")

    #Date Range Filter Method    
    def _date_filter(self, months_back):
        self.browser.click_button_when_visible("//button[@class='css-p5555t']")
        self.browser.wait_until_element_is_visible("//ul[@class='css-vh8n2n']")
        self.browser.click_button_when_visible("//button[@value='Specific Dates']")
        # Find the input fields for start and end date
        start_date_field = "//input[@aria-label='start date']"
        end_date_field = "//input[@aria-label='end date']"
        # Calculate the date range
        end_date = datetime.date.today()
        start_date = end_date - relativedelta(months=months_back)
        # Fill in the start and end date fields
        self.browser.input_text_when_element_is_visible(start_date_field, start_date.strftime("%m/%d/%Y"))
        self.browser.input_text_when_element_is_visible(end_date_field, end_date.strftime("%m/%d/%Y"))
        self.browser.find_element(end_date_field).send_keys(Keys.ENTER)
        print(f"Chosen Date Range from {start_date} to {end_date}")
    
    #Method to show all the news pressing show more button            
    def _press_show_more(self):
        try:
            while self.browser.is_element_visible("//button[normalize-space()='Show More']"):
                self.browser.wait_until_element_is_visible("//button[normalize-space()='Show More']", timeout= 10)
                self.browser.click_button_when_visible("//button[normalize-space()='Show More']")
                time.sleep(2)
        except Exception as ex:
            raise Exception(f"Failed press show more. {ex}")
    
    #Method to filter with section/category
    def _section_filter(self, sections:list) :
        self.browser.click_button_when_visible("//div[@data-testid='section']//button[@type='button']")
        self.browser.wait_until_element_is_visible("//ul[@class='css-64f9ga']")
        section_elements = self.browser.find_elements("//ul[@class='css-64f9ga']//span[@class='css-16eo56s']")
        time.sleep(2)
        for section_element in section_elements:
            for section in sections:
                if section in section_element.text:
                    section_element.click()
        self.browser.click_button_when_visible("//div[@data-testid='section']//button[@type='button']")
        
    #Applying All the Filters 
    def apply_filters(self):
        print("Applying all the filters...")
        #Sort by newest
        self.browser.select_from_list_by_value("css=.css-v7it2b", "newest")
        #Section Filter
        self._section_filter(Config().SECTION)
        #Data Filter
        self._date_filter(Config().MONTHS_BACK)
        #After manually entering the date range, the page loaded an incorrect combination of hours which was resolved after refreshing
        self.browser.reload_page()
        time.sleep(3)
        
    #Show all the news pressing "Show More" button
        self._press_show_more()
       
    #Method to return the news' title
    def _get_title(self, i: int):
        try:
            title_element = self.browser.find_element(f"(//h4[@class='css-2fgx4k'])[{i}]")
            return title_element.text 
        except:
            return "No title found"
    #Method to return the news' description
    def _get_description(self, i: int):
        try:
            description_element = self.browser.find_element(f"//li[@data-testid='search-bodega-result'][{i}]//p[@class='css-16nhkrn']")
            return description_element.text 
        except:    
            return "No description found"
    
    #Method to return the news' date
    def _get_date(self, i: int):
        try:
            date_element = self.browser.find_element(f"(//span[@class='css-17ubb9w'])[{i}]")
            return date_element.text 
        except:
            return "No date found"
    
    #Method to return the news' picture link        
    def _get_picture_link(self, i: int):
        try:
            picture_element = self.browser.find_element(f"//li[@data-testid='search-bodega-result'][{i}]//img")
            return picture_element.get_attribute("src")
        except:
            return "No date found" 
    
    #Method to  download the news' picture
    def _download_picture(self, picture_link, i: int):
        if not os.path.exists("output"):
            os.makedirs("output")
        picture = urllib.request.urlopen(picture_link).read()
        picture_filename = f"news_picture_{i}"
        with open(os.path.join(self.cwd,"output", f"{picture_filename}.jpg"), "wb") as f:
            f.write(picture)
        return picture_filename
    
    #Method to evaluate if the title/description has any ammount of money
    def _contains_money_(self,str1, str2):
        #Pattern $11.1 | $111,111.11 
        pattern1 = r'\$[0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{1,2})?'
        #Pattern 11 dollars | 11 USD
        pattern2 = r'[0-9]+ (?:dollars|USD)'

        match1 = re.search(pattern1, str1)
        match2 = re.search(pattern2, str1)
        match3 = re.search(pattern1, str2)
        match4 = re.search(pattern2, str2)

        return match1 is not None or match2 is not None or match3 is not None or match4 is not None
    
    #Method to get the news' data
    def get_news_data(self):
        self.browser.wait_until_element_is_visible("//div[@class='css-1wa7u5r']")
        news_list = []
        print("Getting data from all the news...")
        news_elements = self.browser.find_elements("//li[@data-testid='search-bodega-result']")
        for i, _ in enumerate(news_elements,start=1):
            try:
                new_info = Info()
                new_info.title = self._get_title(i)
                new_info.date = self._get_date(i)
                new_info.description = self._get_description(i)
                new_info.phrase_count = new_info.title.lower().count(Config.SEARCH_PHRASE.lower()) + new_info.description.lower().count(Config.SEARCH_PHRASE.lower()) 
                new_info.contains_money = self._contains_money_(new_info.title, new_info.description)
                new_info.picture_link = self._get_picture_link(i)
                new_info.picture_filename = self._download_picture(new_info.picture_link, i)
                news_list.append(new_info)
            except:
                print(f"Error in news number {i}")
                continue
        return news_list
    
    
    #Method to save the data in an excel file
    def save_to_excel(self, news_list):
        print("Saving to excel in \output ...")
        wb = openpyxl.Workbook()
        ws = wb.active
  
        ws["A1"] = "Title"
        ws["B1"] = "Date"
        ws["C1"] = "Description"
        ws["D1"] = "Picture Link"
        ws["E1"] = "Picture Filename"
        ws["F1"] = "Search Phrase Count"
        ws["G1"] = "Contains Money"

        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)
        ws["D1"].font = Font(bold=True)
        ws["E1"].font = Font(bold=True)
        ws["F1"].font = Font(bold=True)
        ws["G1"].font = Font(bold=True)


        for i, news in enumerate(news_list, start=2):
            ws[f"A{i}"] = news.title
            ws[f"B{i}"] = news.date
            ws[f"C{i}"] = news.description
            ws[f"D{i}"] = news.picture_link
            ws[f"E{i}"] = news.picture_filename
            ws[f"F{i}"] = news.phrase_count
            ws[f"G{i}"] = news.contains_money

        wb.save(os.path.join(self.cwd, "output", "News.xlsx"))